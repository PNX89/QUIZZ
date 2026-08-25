"""Capture the as-of corpus from the Philadelphia Fed's real-time data set.

RUN THIS BY HAND, NEVER IN CONTINUOUS INTEGRATION. It is the only code here that reaches the
network. Everything the tests and the demonstration use is committed under `data/vintages/`, so
a reader clones the repository and runs it offline with nothing to configure.

    uv sync --group capture
    uv run python scripts/capture_vintages.py

WHY THIS SOURCE AND NOT THE OBVIOUS ONE. The obvious archive of revised economic data is ALFRED,
and it is ruled out by its own terms of use rather than by preference. The FRED terms prohibit
"any data mining, mirroring, robots, scraping, or similar data-gathering or extraction methods
except as expressly allowed by the terms of use applicable to the FRED API", and then prohibit
outright any use of that content "in connection with the development or training of any software
program or system or machine learning, including, but not limited to, large language models".
This repository is a harness for a language model reading economic data, so that second clause is
not a grey area to be argued around. It is the whole of what this repository does.

The Philadelphia Fed publishes the Real-Time Data Set for Macroeconomists for exactly the
question this repository asks: what did the published statistic say at the time somebody acted on
it. Each file below is a single published spreadsheet, downloaded whole rather than harvested,
holding one column per vintage. The underlying series are Bureau of Economic Analysis and Bureau
of Labor Statistics output, which are works of the United States government.

WHAT IS COMMITTED IS AN EXTRACT AND NOT THE DATA SET. Only the declared observation and vintage
windows below, and within them only the rows where a value CHANGED, which is what a revision log
is. The source, the retrieval date and both windows are written to `SOURCE.json` beside the data,
so a reader can go back to the publisher rather than take this repository's word for anything.

THE TRAP THIS SCRIPT EXISTS TO NOT FALL INTO. The first attempt at this corpus used the FRED
graph endpoint with a `vintage_date` parameter on it. That endpoint returns 200 and clean CSV and
silently ignores the parameter: four different vintages of one quarter of real output came back
byte identical, at today's value. A corpus built that way would have held four files named for
four knowing-times, every one of them carrying the latest revision, and every as-of test would
have passed while measuring nothing at all. `assert_revisions_present` below is the guard, and it
is the reason this script fails rather than writes when a publisher's answer stops varying.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import pathlib
import re
import sys
import urllib.request
from dataclasses import dataclass

REPO = pathlib.Path(__file__).resolve().parent.parent
# Inside the package rather than beside it, so an installed wheel carries the corpus and a
# reader who pip installs this gets a working demonstration rather than an import error.
OUT = REPO / "src" / "quizz" / "data" / "vintages"

#: Only vintages from this month onward, and only observations from this period onward.
FIRST_VINTAGE = "2019-01"
FIRST_OBSERVATION = "2015-01"

PUBLISHER = "Federal Reserve Bank of Philadelphia, Real-Time Data Set for Macroeconomists"
PUBLISHER_URL = "https://www.philadelphiafed.org/surveys-and-data/real-time-data-research"
BASE = "https://www.philadelphiafed.org/-/media/FRBP/Assets/Surveys-And-Data/real-time-data/data-files/xlsx/"


@dataclass(frozen=True)
class Series:
    """One published vintage spreadsheet, and what its numbers mean.

    `file` carries the publisher's own cache-busting hash. It is pinned here rather than
    discovered by reading their pages, so this script asks for exactly four known files and
    cannot wander into being a crawler.

    `revisable` is a claim about the statistic and it is checked in BOTH directions. The
    national accounts are estimated from incomplete source data and restated for years. A
    published consumer price index level is not: it is what it was, and the next month's
    release does not reach back and change it. A corpus needs both, because an agent that
    only ever meets revisable numbers learns to hedge about everything.
    """

    name: str
    label: str
    units: str
    file: str
    revisable: bool
    #: The publisher's own first, second and third release figures for the same series, built
    #: by them from the same vintages and expressed as annualised quarterly growth rather than
    #: levels. Captured where it exists, because a second artefact computed a different way is
    #: the only thing that can catch an extraction that is wrong in a self consistent way.
    releases: str | None = None

    @property
    def url(self) -> str:
        return BASE + self.file

    @property
    def releases_url(self) -> str | None:
        return BASE + self.releases if self.releases else None


SERIES = (
    Series(
        "ROUTPUT",
        "Real GNP/GDP",
        "billions of chained dollars, seasonally adjusted",
        "routputMvQd.xlsx?sc_lang=en&hash=B3AD0AED9E268AA077E8D9B8AAF2AC83",
        revisable=True,
        releases="routput_first_second_third.xlsx?sc_lang=en&hash=AB8BB59BBBF6840DE1448851E90D7A80",
    ),
    Series(
        "NOUTPUT",
        "Nominal GNP/GDP",
        "billions of dollars, seasonally adjusted",
        "noutputMvQd.xlsx?sc_lang=en&hash=9A20CEF072FD750998CD62A38E6A8232",
        revisable=True,
    ),
    Series(
        "P",
        "GDP price index",
        "index, seasonally adjusted",
        "pMvQd.xlsx?sc_lang=en&hash=A9B06A9E6E7E92CA131A04BD5AE6856D",
        revisable=True,
    ),
    # Revisable, but by a different mechanism from the three above, and that is why it is here.
    # The national accounts are restated because better source data arrived. This one is
    # restated because the seasonal factors were recalculated: measured over this extract, 474
    # of its value changes land in a FEBRUARY vintage, which is when the annual recalculation
    # is published and rewrites several years of seasonally adjusted history in one step.
    # Nobody learned anything new about the month in question. The number changed anyway.
    Series(
        "PCPI",
        "Consumer price index, seasonally adjusted",
        "index, 1982 to 1984 equals 100",
        "pcpiMvMd.xlsx?sc_lang=en&hash=E41A743DC6423F950B10C3DE7A4F674D",
        revisable=True,
    ),
)

_VINTAGE = re.compile(r"^[A-Za-z_]+(\d{2})M(\d{1,2})$")
_QUARTER = re.compile(r"^(\d{4}):Q([1-4])$")
# Two spellings, because the publisher uses both: quarterly files date a row `2024:Q2` and
# monthly files date it `2024:07`. The first capture read only a `:M` form that nothing uses,
# and the revision guard below caught it as an extract in which nothing had ever been revised,
# which is what a total parse failure looks like from the outside.
_MONTH = re.compile(r"^(\d{4}):M?(\d{1,2})$")


def vintage_month(header: str) -> str | None:
    """`ROUTPUT24M8` becomes `2024-08`, which sorts as a string and reads as a date.

    Returns None for anything else, so a header this does not understand is skipped rather
    than guessed at. The publisher's two digit year is expanded against 1900 and 2000 by the
    only rule that fits a data set starting in 1965: 65 and up is the twentieth century.
    """
    match = _VINTAGE.match(header.strip())
    if not match:
        return None
    year, month = int(match.group(1)), int(match.group(2))
    return f"{1900 + year if year >= 65 else 2000 + year}-{month:02d}"


def observation_period(cell: object) -> str | None:
    """`2024:Q2` stays a quarter; `2024:M7` becomes `2024-07`. Anything else is skipped."""
    text = str(cell).strip() if cell is not None else ""
    quarter = _QUARTER.match(text)
    if quarter:
        return f"{quarter.group(1)}-Q{quarter.group(2)}"
    month = _MONTH.match(text)
    if month:
        return f"{month.group(1)}-{int(month.group(2)):02d}"
    return None


def sort_key(period: str) -> tuple[int, int]:
    """Order `2024-Q2` and `2024-07` alike, so the two frequencies share one comparison."""
    year, rest = period.split("-", 1)
    return (int(year), int(rest[1:]) * 3 if rest.startswith("Q") else int(rest))


def revision_log(rows: list[tuple[object, ...]]) -> list[tuple[str, str, float]]:
    """Every point where a value CHANGED, which is the whole of what a vintage file says.

    Storing one row per vintage per observation would be storing the same number ninety times
    and would be redistributing the data set rather than extracting from it. What is kept is
    the first vintage carrying each distinct value, which is enough to answer any as-of
    question by looking for the newest row at or before a knowing-time.

    A cell that is empty or `#N/A` means the observation had not been published in that
    vintage yet. That absence is information rather than a gap: it is the difference between
    an answer and "that number did not exist when you are asking about".
    """
    header, *body = rows
    vintages = [(index, vintage_month(str(cell))) for index, cell in enumerate(header)]
    keep = [(index, month) for index, month in vintages if month and month >= FIRST_VINTAGE]
    out: list[tuple[str, str, float]] = []
    for row in body:
        period = observation_period(row[0] if row else None)
        if not period or sort_key(period) < sort_key(FIRST_OBSERVATION):
            continue
        previous: float | None = None
        for index, month in keep:
            raw = row[index] if index < len(row) else None
            if raw is None or isinstance(raw, str):
                continue  # not published in this vintage, or the publisher's #N/A marker
            value = round(float(raw), 6)
            if previous is None or value != previous:
                out.append((period, month, value))
                previous = value
    return sorted(out, key=lambda r: (sort_key(r[0]), r[1]))


def assert_revision_behaviour(series: Series, log: list[tuple[str, str, float]]) -> None:
    """Check the declared revision behaviour, in both directions, and refuse to write if it fails.

    The failure this started as: a source that answers every knowing-time with today's number
    produces a corpus where each observation appears exactly once. That looks like a clean data
    set and is a broken capture, and this caught exactly that twice, once for the ignored
    vintage parameter that made this file necessary and once for a date pattern of my own that
    matched nothing at all.

    The other direction matters just as much. A series declared unrevisable that suddenly shows
    revisions means either the publisher changed the statistic or this script is reading the
    wrong file, and both are worth stopping for rather than absorbing.
    """
    counts: dict[str, int] = {}
    for period, _, _ in log:
        counts[period] = counts.get(period, 0) + 1
    revised = sorted(period for period, count in counts.items() if count > 1)
    if series.revisable and len(revised) < 2:
        raise SystemExit(
            f"{series.name} is declared revisable and only {len(revised)} observations in this "
            "extract were ever revised. A vintage source answering every knowing-time with the "
            "same number is a broken capture, not a clean data set. Refusing to write it."
        )
    if not series.revisable and revised:
        raise SystemExit(
            f"{series.name} is declared unrevisable and {len(revised)} observations were revised, "
            f"the first being {revised[0]}. Either the publisher restated a statistic that is not "
            "supposed to be restated, or this is reading the wrong file. Refusing to write it."
        )


def release_table(raw: bytes) -> list[tuple[str, float, float, float]]:
    """The publisher's own first, second and third release growth rates, one row per quarter.

    Rows where any of the three is the publisher's `#N/A` marker are dropped rather than
    filled: a release that has not happened yet has no figure, and inventing one would be the
    same class of mistake as answering a question about a period that was not published.
    """
    import io

    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    out: list[tuple[str, float, float, float]] = []
    for row in workbook["DATA"].iter_rows(values_only=True):
        label = str(row[0]) if row and row[0] else ""
        if not _QUARTER.match(label.strip()):
            continue
        three = row[1:4]
        if any(not isinstance(cell, (int, float)) for cell in three):
            continue
        period = observation_period(label)
        assert period is not None
        first, second, third = (float(cell) for cell in three)  # type: ignore[arg-type]
        if period >= FIRST_OBSERVATION:
            out.append((period, first, second, third))
    return out


def fetch(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"Accept": "*/*"})
    with urllib.request.urlopen(request, timeout=120) as response:
        return bytes(response.read())


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="fetch and report, write nothing")
    args = parser.parse_args(argv)

    try:
        import openpyxl
    except ModuleNotFoundError:
        print("openpyxl is a capture dependency: uv sync --group capture", file=sys.stderr)
        return 2

    OUT.mkdir(parents=True, exist_ok=True)
    provenance: dict[str, object] = {
        "publisher": PUBLISHER,
        "publisher_url": PUBLISHER_URL,
        "retrieved": dt.date.today().isoformat(),
        "first_vintage": FIRST_VINTAGE,
        "first_observation": FIRST_OBSERVATION,
        "note": (
            "An extract, not the data set. Only the windows above, and within them only the "
            "vintages where a value changed. Go to the publisher for the whole of it."
        ),
        "series": [],
    }

    for series in SERIES:
        workbook = openpyxl.load_workbook(
            __import__("io").BytesIO(fetch(series.url)), read_only=True, data_only=True
        )
        sheet = workbook[workbook.sheetnames[0]]
        log = revision_log([tuple(row) for row in sheet.iter_rows(values_only=True)])
        assert_revision_behaviour(series, log)

        path = OUT / f"{series.name}.csv"
        revised = len(
            [1 for period in {r[0] for r in log} if len([r for r in log if r[0] == period]) > 1]
        )
        print(
            f"{series.name}: {len(log)} rows, {len({r[0] for r in log})} observations, "
            f"{revised} of them revised at least once"
        )
        if not args.dry_run:
            with path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                writer.writerow(["observation", "vintage", "value"])
                writer.writerows(log)
        entry = {
            "name": series.name,
            "label": series.label,
            "units": series.units,
            "url": series.url,
            "revisable": series.revisable,
            "rows": len(log),
            "observations": len({r[0] for r in log}),
            "vintages": len({r[1] for r in log}),
        }
        if series.releases_url:
            releases = release_table(fetch(series.releases_url))
            if not args.dry_run:
                with (OUT / f"{series.name}_releases.csv").open(
                    "w", newline="", encoding="utf-8"
                ) as handle:
                    writer = csv.writer(handle)
                    writer.writerow(["observation", "first", "second", "third"])
                    writer.writerows(releases)
            entry["releases_url"] = series.releases_url
            entry["release_rows"] = len(releases)
            print(f"{series.name}: {len(releases)} published first, second and third releases")

        assert isinstance(provenance["series"], list)
        provenance["series"].append(entry)

    if not args.dry_run:
        (OUT / "SOURCE.json").write_text(json.dumps(provenance, indent=2) + "\n", encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
